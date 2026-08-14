"""
Assembler & Output Reconstructor Sub-Agent.
Unmasks tags and reconstructs the target localized file matching
the exact original format (JSON, CSV, XLSX, RPY, Subtitles, DOCX, MD, TXT).
"""

import os
import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from ..models import TranslationItem
from ..tag_protector import TagProtector


class AssemblerAgent:
    """Sub-agent responsible for unmasking tags and writing translated files."""

    @classmethod
    def assemble_results(
        cls,
        items: List[TranslationItem],
        format_type: str,
        file_context: Dict[str, Any],
        output_path: Optional[Union[str, Path]] = None,
    ) -> str:
        """
        Unmasks tags for all items and writes output file or returns reconstructed text.
        """
        # 1. Unmask tags for every item
        for item in items:
            raw_trans = item.translated_text if item.translated_text is not None else item.source_text
            item.translated_text = TagProtector.unmask_text(raw_trans, item.tag_map)

        # 2. Reconstruct based on format
        if format_type == "json":
            result_str = cls._assemble_json(items, file_context)
        elif format_type == "csv":
            result_str = cls._assemble_csv(items, file_context)
        elif format_type == "xlsx":
            return cls._assemble_xlsx(items, file_context, output_path)
        elif format_type == "renpy":
            result_str = cls._assemble_renpy(items, file_context)
        elif format_type == "subtitle":
            result_str = cls._assemble_subtitles(items, file_context)
        elif format_type == "po":
            result_str = cls._assemble_po(items, file_context)
        elif format_type == "docx":
            return cls._assemble_docx(items, file_context, output_path)
        elif format_type == "markdown":
            result_str = cls._assemble_markdown(items, file_context)
        else:
            # Default TXT
            result_str = cls._assemble_txt(items)

        # 3. Write to disk if output_path specified
        if output_path:
            out_file = Path(output_path)
            out_file.parent.mkdir(parents=True, exist_ok=True)
            out_file.write_text(result_str, encoding="utf-8")
            return str(out_file)

        return result_str

    # -------------------------------------------------------------
    # Format Reconstructors
    # -------------------------------------------------------------

    @classmethod
    def _assemble_json(cls, items: List[TranslationItem], file_context: Dict[str, Any]) -> str:
        raw_data = file_context.get("raw_data")
        item_map = {item.id: item.translated_text for item in items}

        if isinstance(raw_data, dict):
            new_dict = {}
            for k, v in raw_data.items():
                str_k = str(k)
                if isinstance(v, str):
                    new_dict[k] = item_map.get(str_k, v)
                elif isinstance(v, dict):
                    # Mtool format: {"src": "...", "trans": "..."}
                    new_entry = dict(v)
                    trans_text = item_map.get(str_k, "")
                    if "trans" in new_entry or "src" in new_entry:
                        new_entry["trans"] = trans_text
                    elif "message" in new_entry:
                        new_entry["message"] = trans_text
                    elif "text" in new_entry:
                        new_entry["text"] = trans_text
                    new_dict[k] = new_entry
                else:
                    new_dict[k] = v
            return json.dumps(new_dict, ensure_ascii=False, indent=2)

        elif isinstance(raw_data, list):
            new_list = []
            for idx, entry in enumerate(raw_data):
                str_idx = str(idx)
                if isinstance(entry, str):
                    new_list.append(item_map.get(str_idx, entry))
                elif isinstance(entry, dict):
                    item_id = str(entry.get("id", idx))
                    new_entry = dict(entry)
                    trans_text = item_map.get(item_id, "")
                    if "trans" in new_entry or "src" in new_entry:
                        new_entry["trans"] = trans_text
                    elif "target" in new_entry:
                        new_entry["target"] = trans_text
                    elif "text" in new_entry:
                        new_entry["text"] = trans_text
                    new_list.append(new_entry)
                else:
                    new_list.append(entry)
            return json.dumps(new_list, ensure_ascii=False, indent=2)

        # Fallback flat dict
        return json.dumps({it.id: it.translated_text for it in items}, ensure_ascii=False, indent=2)

    @classmethod
    def _assemble_csv(cls, items: List[TranslationItem], file_context: Dict[str, Any]) -> str:
        delimiter = file_context.get("delimiter", ",")
        header = file_context.get("header", ["id", "source", "target"])
        rows = file_context.get("rows", [])

        item_lookup = {it.raw_metadata.get("row_idx"): it.translated_text for it in items}

        out_rows = []
        if rows:
            # Reconstruct with existing rows
            tgt_col = None
            for it in items:
                if it.raw_metadata.get("tgt_col") is not None:
                    tgt_col = it.raw_metadata.get("tgt_col")
                    break

            new_header = list(header)
            if tgt_col is None:
                new_header.append("target")
                tgt_col = len(new_header) - 1

            out_rows.append(new_header)
            for r_idx, row in enumerate(rows[1:], start=1):
                new_row = list(row)
                while len(new_row) <= tgt_col:
                    new_row.append("")
                if r_idx in item_lookup and item_lookup[r_idx]:
                    new_row[tgt_col] = item_lookup[r_idx]
                out_rows.append(new_row)
        else:
            out_rows.append(["id", "speaker", "source", "target"])
            for it in items:
                out_rows.append([it.id, it.speaker or "", it.source_text, it.translated_text or ""])

        import io
        output = io.StringIO()
        writer = csv.writer(output, delimiter=delimiter)
        writer.writerows(out_rows)
        return output.getvalue()

    @classmethod
    def _assemble_xlsx(cls, items: List[TranslationItem], file_context: Dict[str, Any], output_path: Optional[Union[str, Path]]) -> str:
        import openpyxl

        src_file = file_context.get("file_path")
        if src_file and Path(src_file).exists():
            wb = openpyxl.load_workbook(src_file)
            sheet = wb.active
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["id", "speaker", "source", "target"])

        item_lookup = {it.raw_metadata.get("row_idx"): it.translated_text for it in items}

        for it in items:
            r_idx = it.raw_metadata.get("row_idx")
            tgt_col = it.raw_metadata.get("tgt_col")
            if r_idx is not None:
                col_idx = (tgt_col + 1) if (tgt_col is not None) else 4
                sheet.cell(row=r_idx, column=col_idx, value=it.translated_text or "")

        target_out = output_path or (Path(src_file).with_name(f"{Path(src_file).stem}_translated.xlsx") if src_file else "output_translated.xlsx")
        wb.save(target_out)
        return str(target_out)

    @classmethod
    def _assemble_renpy(cls, items: List[TranslationItem], file_context: Dict[str, Any]) -> str:
        lines = list(file_context.get("lines", []))
        item_lookup = {it.raw_metadata.get("line_idx"): it for it in items}

        for idx, line in enumerate(lines):
            if idx in item_lookup:
                it = item_lookup[idx]
                trans = (it.translated_text or "").replace('"', r'\"')
                if it.raw_metadata.get("type") == "dialogue":
                    if it.speaker:
                        lines[idx] = f'    {it.speaker} "{trans}"'
                    else:
                        lines[idx] = f'    "{trans}"'
                elif it.raw_metadata.get("type") == "old_string":
                    lines[idx] = f'    new "{trans}"'

        return "\n".join(lines)

    @classmethod
    def _assemble_subtitles(cls, items: List[TranslationItem], file_context: Dict[str, Any]) -> str:
        ext = file_context.get("ext", ".srt")
        if ext == ".srt":
            blocks = []
            for it in items:
                idx = it.raw_metadata.get("index", it.id)
                tc = it.raw_metadata.get("timecode", "00:00:00,000 --> 00:00:05,000")
                text = it.translated_text or it.source_text
                blocks.append(f"{idx}\n{tc}\n{text}\n")
            return "\n".join(blocks)
        elif ext == ".ass":
            lines = list(file_context.get("raw_lines", []))
            item_lookup = {it.raw_metadata.get("line_idx"): it.translated_text for it in items}
            for idx, line in enumerate(lines):
                if idx in item_lookup:
                    prefix = items[0].raw_metadata.get("prefix", "Dialogue: 0,0:00:00.00,0:00:05.00,Default,,0,0,0,,")
                    lines[idx] = f"{prefix}{item_lookup[idx]}"
            return "\n".join(lines)
        else:
            return "\n".join(it.translated_text or it.source_text for it in items)

    @classmethod
    def _assemble_po(cls, items: List[TranslationItem], file_context: Dict[str, Any]) -> str:
        raw_text = file_context.get("raw_text", "")
        # Reconstruct PO with msgstr replaced
        out_lines = []
        for it in items:
            src = it.source_text.replace('"', r'\"').replace('\n', r'\n')
            tgt = (it.translated_text or "").replace('"', r'\"').replace('\n', r'\n')
            out_lines.append(f'msgid "{src}"\nmsgstr "{tgt}"\n')
        return "\n".join(out_lines)

    @classmethod
    def _assemble_docx(cls, items: List[TranslationItem], file_context: Dict[str, Any], output_path: Optional[Union[str, Path]]) -> str:
        import docx

        doc = docx.Document()
        for it in items:
            text = it.translated_text or it.source_text
            doc.add_paragraph(text)

        target_out = output_path or "translated_document.docx"
        doc.save(target_out)
        return str(target_out)

    @classmethod
    def _assemble_markdown(cls, items: List[TranslationItem], file_context: Dict[str, Any]) -> str:
        return "\n\n".join(it.translated_text or it.source_text for it in items)

    @classmethod
    def _assemble_txt(cls, items: List[TranslationItem]) -> str:
        lines = []
        for it in items:
            text = it.translated_text or it.source_text
            if it.speaker:
                lines.append(f"{it.speaker}: {text}")
            else:
                lines.append(text)
        return "\n".join(lines)
